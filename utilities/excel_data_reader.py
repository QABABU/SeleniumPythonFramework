import openpyxl


class DataReader:

    def __init__(self, file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        self.sheet = workbook[sheet_name]

    def get_data_sheet(self):
        return self.sheet

    def get_row_count(self):
        return self.sheet.max_row

    def get_cell_data(self, row_num, col_num):
        value = self.sheet.cell(row=row_num, column=col_num).value
        return value

    def get_current_test_row(self, test_name):
        current_test_row_num = -1
        row_count = self.get_row_count()
        for r in range(1, row_count + 1):
            if self.sheet.cell(row=r, column=1).value == test_name:
                current_test_row_num = r
                break
        return current_test_row_num

    def get_data_sets_count(self, test_name):
        current_test_row_num = self.get_current_test_row(test_name)
        print("---- current_test_row_num ---- ", current_test_row_num)
        sets = 0
        if current_test_row_num != -1:
            first_data_set_row_num = current_test_row_num + 2
            print("---- first_data_set_row_num ---- ", first_data_set_row_num)
            while self.get_cell_data(first_data_set_row_num, 1) is not None:
                sets += 1
                first_data_set_row_num += 1
            print("---- total_data_sets ---- ", sets)
            return sets
        else:
            print("No such test exist in the test data sheet")

    def get_test_data(self, test_name):
        current_test_row_num = self.get_current_test_row(test_name)
        sets = self.get_data_sets_count(test_name)
        if current_test_row_num != -1 and sets != 0:
            first_data_set_row_num = current_test_row_num + 2
            last_data_set_row_num = first_data_set_row_num + sets - 1
            all_rows = []
            for rows in self.sheet.iter_rows(first_data_set_row_num, last_data_set_row_num):
                single_row = []
                if rows[0].value is 'Y':
                    for cell in rows:
                        cell_value = cell.value
                        if cell_value != 'Y' and cell_value is not None:
                            single_row.append(cell_value)
                            print(cell_value)
                    all_rows.append(tuple(single_row))
            return all_rows
        else:
            print("No such test exist in the test data sheet")
